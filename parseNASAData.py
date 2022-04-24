from posixpath import split
from urllib import response
import requests
from bs4 import BeautifulSoup as bs
import re
from openpyxl import Workbook
from openpyxl.styles import Font 
from openpyxl.drawing.image import Image
from matplotlib import pyplot as plt
import pandas as pd 
import tk

#######################################################################################

URL = 'https://mars.nasa.gov/technology/helicopter'
FLIGHT_LOG_TABLE = 'flight-log-table'
TAG_RE = re.compile(r'<[^>]+>')
HEADERS = ('flight:',
           'sol:',
           'date:',
           'hor_dist (m):',
           'hor_dist (ft):',
           'max_alt (m):',
           'max_alt (ft):',
           'max_gnd_speed (m/s):',
           'max_gnd_speed (mph):',
           'duration:',
           'route:')
EXCEL_WB = "./MyExcelWorkbook.xlsx"
IMAGE_HEIGHT = 500
IMAGE_WIDTH = 1000

wb = Workbook()
sheet = wb.active
sheet.Name = "Flight Data"
    
#######################################################################################
### Functions ###
#######################################################################################

def remove_tags(text):
    # return ','.join(xml.etree.ElementTree.fromstring(text).itertext())
    return TAG_RE.sub('|', text)


def print_flight_info_to_stdout(flight_info):

    # print("flight:\tsol:\tdate:\thor_dist(m):\thor_dist(ft):\tmax_alt(m):\tmax_alt(ft):\tmax_gnd_speed(m/s):\tmax_gnd_speed(mph):\tduration:\troute:")

    print("\n")
    for flight in flight_info:
        print("%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s" %
              (flight['flight'],
               flight['sol'],
               flight['date'],
               flight['hor_dist']['m'],
               flight['hor_dist']['ft'],
               flight['max_alt']['m'],
               flight['max_alt']['ft'],
               flight['max_gnd_speed']['m/s'],
               flight['max_gnd_speed']['mph'],
               flight['duration'],
               flight['route']
               ))
    print("\n")

def print_flight_info_to_excel(flight_info):
    
    column = 1 
    for header in HEADERS:
        sheet.cell(row = 1, column = column).font = Font( bold = True)
        sheet.cell(row = 1, column = column).value = header.upper()
        column = column + 1
        
    row = 2
    for flight in flight_info:
    
        sheet.cell(row = row, column = 1).value = flight['flight']
        sheet.cell(row = row, column = 2).value = flight['sol']
        sheet.cell(row = row, column = 3).value = flight['date']
        sheet.cell(row = row, column = 4).value = flight['hor_dist']['m']
        sheet.cell(row = row, column = 5).value = flight['hor_dist']['ft']
        sheet.cell(row = row, column = 6).value = flight['max_alt']['m']
        sheet.cell(row = row, column = 7).value = flight['max_alt']['ft']
        sheet.cell(row = row, column = 8).value = flight['max_gnd_speed']['m/s']
        sheet.cell(row = row, column = 9).value = flight['max_gnd_speed']['mph']
        sheet.cell(row = row, column = 10).value = flight['duration']
        sheet.cell(row = row, column = 11).value =  flight['route']        
        row = row + 1

    wb.save(filename=EXCEL_WB)

def print_flight_info_to_plots_1(data, plot_name):
    
   plt.style.use('seaborn')
   plt.subplot()
   
   plt.plot(data['flights'], data['hor_dist_in_m'], label='Horizontal Distance (m)')
   plt.plot(data['flights'], data['max_alt_in_m'], label='Maximum Altitude (m)')
   plt.plot(data['flights'], data['max_gnd_speed_in_m_per_s'], label='Maximun Ground Speed (m/s)')
   plt.plot(data['flights'], data['duration'], label='Duration (sec)')
   
   plt.legend()
   
   plt.title('Flight Data')
   plt.xlabel('Flight Number')
   # plt.ylabel('Y-axis: TBD')
   
   plt.tight_layout()
   plt.autoscale(enable=True, axis='y')
   
   plt.savefig(plot_name)
   plt.show()
   plt.close()

   
def print_flight_info_to_plots_2(data, plot_name):
   
   plt.style.use('seaborn')
   fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1)
   
   ax1.plot(data['flights'], data['hor_dist_in_m'], label='Horizontal Distance (m)')
   ax2.plot(data['flights'], data['max_alt_in_m'], label='Maximum Altitude (m)')
   ax2.plot(data['flights'], data['max_gnd_speed_in_m_per_s'], label='Maximun Ground Speed (m/s)')
   ax1.plot(data['flights'], data['duration'], label='Duration (sec)')
   
   ax1.legend()
   ax2.legend()
   
   ax1.set_title('Flight Data')
   ax1.set_xlabel('Flight Number')
   # ax1.set_ylabel('Y-axis: TBD')
   
   ax2.set_title('Flight Data')
   ax2.set_xlabel('Flight Number')
   # ax2.set_ylabel('Y-axis: TBD')
   
   plt.tight_layout()
   plt.autoscale(enable=True, axis='y')
   
   plt.savefig(plot_name)
   plt.show()
   plt.close()

def print_flight_info_to_plots_4(data, plot_name):
       
   plt.style.use('seaborn')

   fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=2)
   
   ax1[0].plot(data['flights'], data['hor_dist_in_m'], label='Horizontal Distance (m)')
   ax1[1].plot(data['flights'], data['max_alt_in_m'], label='Maximum Altitude (m)')
   ax2[0].plot(data['flights'], data['max_gnd_speed_in_m_per_s'], label='Maximun Ground Speed (m/s)')
   ax2[1].plot(data['flights'], data['duration'], label='Duration (sec)')
   
   ax1[0].legend()
   ax1[1].legend()
   ax2[0].legend()
   ax2[1].legend()
   
   ax1[0].set_title('Flight Data')
   ax1[0].set_xlabel('Flight Number')
   # ax1[0].set_ylabel('Y-axis: TBD')
   
   ax1[1].set_title('Flight Data')
   ax1[1].set_xlabel('Flight Number')
   # ax1[1].set_ylabel('Y-axis: TBD')
   
   ax2[0].set_title('Flight Data')
   ax2[0].set_xlabel('Flight Number')
   # ax2[0].set_ylabel('Y-axis: TBD')
   
   ax2[1].set_title('Flight Data')
   ax2[1].set_xlabel('Flight Number')
   # ax2[1].set_ylabel('Y-axis: TBD')
   
   plt.tight_layout()
   plt.autoscale(enable=True, axis='y')
   
   plt.savefig(plot_name)
   plt.show()
   plt.close()
   
   
   
#######################################################################################
### TODO: Part 1: ###
# Pull data from website and add to x-element array (25 elements in this case)
#######################################################################################


response = requests.get(URL)

if response.status_code != 200:
    print("Error. Response code: %d", response.status_code)
    exit()

table = bs(response.content, 'html.parser')
flight_info = str(
    table.find(id=FLIGHT_LOG_TABLE).get_text).replace('\n', '').split('</tr>')

#######################################################################################
### TODO: Part 2 ###
# Strip away html tags so data is easily portable to csv/Excel
#######################################################################################

flight_info_data = []
flights = []
hor_dist_in_m = []
max_alt_in_m = []
max_gnd_speed_in_m_per_s = []
duration = []
   
# Ignore elements 0, 1 and len(flight_info) because those are beginning/end of table tags
# for i in range(2, len(flight_info) - 1):
#     print(flight_info[i])

for i in range(2, len(flight_info) - 1):

    plain_text = remove_tags(flight_info[i]).split("|")
    # print(plain_text)

    temp_dict = {
        "flight": plain_text[2],
        "sol": plain_text[4],
        "date": plain_text[10],
        "hor_dist": {
            "m": plain_text[12],
            "ft": plain_text[14]
        },
        "max_alt": {
            "m": plain_text[16],
            "ft": plain_text[18]
        },
        "max_gnd_speed": {
            "m/s": plain_text[20],
            "mph": plain_text[22]},
        "duration": plain_text[24],
        "route": plain_text[26],
    }

    flight_info_data.append(temp_dict)
    
    ### Arrays that will specifically be used for graphing ### 
    flights.append(plain_text[2])
    hor_dist_in_m.append(plain_text[12])
    max_alt_in_m.append(plain_text[16])
    max_gnd_speed_in_m_per_s.append(plain_text[20])
    duration.append(plain_text[24])

print_flight_info_to_stdout(flight_info_data)

#######################################################################################
### TODO: Part 3 ###
# Build csv file tags using Excel objects
#######################################################################################

print_flight_info_to_excel(flight_info_data)

#######################################################################################
### TODO: Part 4 ### 
# manipulate csv data with graphs, etc ...
#######################################################################################

data = {    
    'flights': list(map(int,flights)), 
    'hor_dist_in_m': list(map(float, hor_dist_in_m)), 
    'max_alt_in_m': list(map(float, max_alt_in_m)), 
    'max_gnd_speed_in_m_per_s': list(map(float, max_gnd_speed_in_m_per_s)), 
    'duration': list(map(float,duration ))
        }

print_flight_info_to_plots_1(data, './flight_data_plot_1.png')
print_flight_info_to_plots_2(data, './flight_data_plot_2.png')
print_flight_info_to_plots_4(data, './flight_data_plot_4.png')

image1 = Image('./flight_data_plot_1.png')
image2 = Image('./flight_data_plot_2.png')
image4 = Image('./flight_data_plot_4.png')

# image1.height = IMAGE_HEIGHT
# image1.width = IMAGE_WIDTH

# image2.height = IMAGE_HEIGHT
# image2.width = IMAGE_WIDTH

# image4.height = IMAGE_HEIGHT
# image4.width = IMAGE_WIDTH

sheet.add_image(image1, "P1")
sheet.add_image(image2, "P31")
sheet.add_image(image4, "P61")

wb.save(filename=EXCEL_WB)
